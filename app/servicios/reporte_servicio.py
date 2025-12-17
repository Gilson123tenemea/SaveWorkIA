# app/servicios/reporte_servicio.py
import io
import json
from datetime import datetime
from typing import Optional, List, Dict, Tuple, Any
from sqlalchemy import or_
from collections import defaultdict

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_

from fastapi import HTTPException

from app.modelos.reporte import Reporte
from app.modelos.registros_asistencia import RegistroAsistencia
from app.modelos.evidencias_fallo import EvidenciaFallo
from app.modelos.zona_modelo import Zona
from app.modelos.trabajador import Trabajador
from app.modelos.persona import Persona
from app.modelos.supervisor import Supervisor
from app.modelos.registrosupervisorinspector import RegistroSupervisorInspector
from app.modelos.empresa_modelo import Empresa
from app.modelos.inspector import Inspector
from app.modelos.inspector_zona import InspectorZona
import json

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# EXCEL
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.modelos.camara_modelo import Camara
from app.modelos.zona_modelo import Zona
from app.modelos.zona_epp import ZonaEpp
from app.utils.mapeo_epp import MAPEO_EPP_YOLO
# -----------------------------
# Helpers
# -----------------------------
def _parse_date(date_str: str) -> datetime:
    """
    Acepta 'YYYY-MM-DD' y lo convierte a datetime (00:00:00).
    """
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail=f"Fecha inválida: {date_str}. Use YYYY-MM-DD.")


def _end_of_day(dt: datetime) -> datetime:
    return dt.replace(hour=23, minute=59, second=59, microsecond=0)


def registrar_reporte(
    db: Session,
    tipo_reporte: str,
    formato: str,
    filtros: Optional[dict],
    generado_por: str,
    id_empresa: int,
    id_inspector: Optional[int] = None,
) -> Reporte:
    rep = Reporte(
        tipo_reporte=tipo_reporte,
        formato=formato,
        filtros=json.dumps(filtros, ensure_ascii=False) if filtros else None,
        generado_por=generado_por,
        id_empresa=id_empresa,
        id_inspector=id_inspector,
        borrado=True,
    )
    db.add(rep)
    db.commit()
    db.refresh(rep)
    return rep


# -----------------------------
# 3.1) ESTADÍSTICAS (BARRAS)
# -----------------------------
def barras_incumplimiento_por_zona(
    db: Session,
    id_inspector: int,
    id_empresa: int,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
) -> List[Tuple[str, int]]:
    """
    Incumplimientos = registros con evidencia_fallo asociada (tu modelo actual).
    """
    filters = [
        RegistroAsistencia.id_inspector == id_inspector,
        RegistroAsistencia.id_empresa == id_empresa,
    ]

    if fecha_desde:
        d = _parse_date(fecha_desde)
        filters.append(RegistroAsistencia.fecha_hora >= d)
    if fecha_hasta:
        h = _end_of_day(_parse_date(fecha_hasta))
        filters.append(RegistroAsistencia.fecha_hora <= h)

    q = (
        db.query(
            Zona.nombreZona.label("zona"),
            func.count(EvidenciaFallo.id_evidencia).label("incumplimientos"),
        )
        .join(RegistroAsistencia, RegistroAsistencia.id_zona == Zona.id_Zona)
        .join(EvidenciaFallo, EvidenciaFallo.id_registro == RegistroAsistencia.id_registro)
        .filter(and_(*filters))
        .group_by(Zona.nombreZona)
        .order_by(func.count(EvidenciaFallo.id_evidencia).desc())
    )

    return [(r.zona, int(r.incumplimientos)) for r in q.all()]


def barras_cumplimiento_por_zona(
    db: Session,
    id_inspector: int,
    id_empresa: int,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
) -> List[Tuple[str, int]]:
    """
    Cumplimientos = registros donde cumple_epp = True
    (independiente de evidencia).
    """
    filters = [
        RegistroAsistencia.id_inspector == id_inspector,
        RegistroAsistencia.id_empresa == id_empresa,
        RegistroAsistencia.cumple_epp == True,  # noqa: E712
    ]

    if fecha_desde:
        d = _parse_date(fecha_desde)
        filters.append(RegistroAsistencia.fecha_hora >= d)
    if fecha_hasta:
        h = _end_of_day(_parse_date(fecha_hasta))
        filters.append(RegistroAsistencia.fecha_hora <= h)

    q = (
        db.query(
            Zona.nombreZona.label("zona"),
            func.count(RegistroAsistencia.id_registro).label("cumplimientos"),
        )
        .join(RegistroAsistencia, RegistroAsistencia.id_zona == Zona.id_Zona)
        .filter(and_(*filters))
        .group_by(Zona.nombreZona)
        .order_by(func.count(RegistroAsistencia.id_registro).desc())
    )

    return [(r.zona, int(r.cumplimientos)) for r in q.all()]


# -----------------------------
# 3.2) ESTADÍSTICAS (PASTEL)
#     "EPP más cumplido"
# -----------------------------
EPP_LABELS = [
    ("Casco", ["casco", "helmet"]),
    ("Chaleco", ["chaleco", "vest"]),
    ("Gafas", ["gafas", "glasses", "goggles"]),
    ("Guantes", ["guantes", "gloves"]),
    ("Botas", ["botas", "boots"]),
]

def pastel_epp_mas_cumplido(
    db: Session,
    id_empresa: int,
    id_inspector: Optional[int] = None,
    id_zona: Optional[int] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
) -> List[Tuple[str, int]]:

    conteo_epp = defaultdict(int)

    # -----------------------------
    # Query base (SIN evidencia_fallo)
    # -----------------------------
    query = (
        db.query(RegistroAsistencia)
        .options(
            joinedload(RegistroAsistencia.camara).joinedload(Camara.zona)
        )
        .filter(RegistroAsistencia.id_empresa == id_empresa)
    )

    if id_inspector:
        query = query.filter(RegistroAsistencia.id_inspector == id_inspector)

    if id_zona:
        query = query.filter(RegistroAsistencia.id_zona == id_zona)

    if fecha_desde:
        query = query.filter(RegistroAsistencia.fecha_hora >= _parse_date(fecha_desde))

    if fecha_hasta:
        query = query.filter(
            RegistroAsistencia.fecha_hora <= _end_of_day(_parse_date(fecha_hasta))
        )

    registros = query.all()

    if not registros:
        return []

    # -----------------------------
    # Lógica de cumplimiento REAL
    # -----------------------------
    for reg in registros:

        # 🔥 Obtener evidencia (como en tus otros servicios)
        evidencia = (
            db.query(EvidenciaFallo)
            .filter(EvidenciaFallo.id_registro == reg.id_registro)
            .first()
        )

        zona = reg.camara.zona

        # 🔹 EPP obligatorios reales de la zona
        epps_zona = obtener_epp_humanos_por_zona(db, zona.id_Zona)

        # 🔹 EPP que FALLARON
        detecciones = []
        if evidencia and evidencia.detalle_fallo:
            detecciones = [
                d.lower().replace("falta", "").strip()
                for d in evidencia.detalle_fallo.split(",")
                if d.strip()
            ]

        # 🔹 Contar cumplimiento
        for epp in epps_zona:
            epp_norm = epp.lower().strip()

            # Si NO está en detecciones → CUMPLIÓ
            if epp_norm not in detecciones:
                conteo_epp[epp_norm] += 1

    # -----------------------------
    # Formato final
    # -----------------------------
    resultados = [
        (epp.capitalize(), total)
        for epp, total in conteo_epp.items()
    ]

    resultados.sort(key=lambda x: x[1], reverse=True)
    return resultados


def obtener_epp_humanos_por_zona(db: Session, id_zona: int) -> list[str]:
    """
    Devuelve los EPP obligatorios y activos de una zona
    SIN duplicados
    Ej: ["casco", "gafas"]
    """

    epps = (
        db.query(ZonaEpp.tipo_epp)
        .filter(
            ZonaEpp.id_zona == id_zona,
            ZonaEpp.activo == True,
            ZonaEpp.obligatorio == True
        )
        .distinct()
        .all()
    )

    # Convierte de [(casco,), (gafas,)] → ["casco", "gafas"]
    return [epp[0] for epp in epps]

# -----------------------------
# 3.3) PDF: Trabajadores por zona y rango (INSPECTOR)
# -----------------------------
def generar_pdf_trabajadores_zona(
    db: Session,
    id_inspector: int,
    id_zona: int,
    fecha_desde: str,
    fecha_hasta: str,
) -> bytes:
    d = _parse_date(fecha_desde)
    h = _end_of_day(_parse_date(fecha_hasta))

    rows = (
        db.query(
            RegistroAsistencia.fecha_hora,
            RegistroAsistencia.cumple_epp,
            Zona.nombreZona,
            Trabajador.codigo_trabajador,
            Persona.cedula,
            Persona.nombre,
            Persona.apellido,
            Persona.correo,
            Persona.telefono,
            EvidenciaFallo.detalle_fallo,
            EvidenciaFallo.observaciones,
        )
        .join(Zona, Zona.id_Zona == RegistroAsistencia.id_zona)
        .join(Trabajador, Trabajador.id_trabajador == RegistroAsistencia.id_trabajador)
        .join(Persona, Persona.id_persona == Trabajador.id_persona_trabajador)
        .outerjoin(EvidenciaFallo, EvidenciaFallo.id_registro == RegistroAsistencia.id_registro)
        .filter(
    RegistroAsistencia.id_inspector == id_inspector,
    RegistroAsistencia.id_zona == id_zona,
    RegistroAsistencia.fecha_hora >= d,
    RegistroAsistencia.fecha_hora <= h,
)

        .order_by(RegistroAsistencia.fecha_hora.desc())
        .all()
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=18,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(
        f"Reporte de Trabajadores – Zona {id_zona}",
        styles["Title"]
    ))
    story.append(Paragraph(
        f"Inspector ID: {id_inspector} | Desde {fecha_desde} hasta {fecha_hasta}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 10))

    data = [[
        "Fecha/Hora", "Zona", "Código", "Cédula",
        "Nombre", "Correo", "Teléfono",
        "Estado EPP", "Detalle", "Observaciones"
    ]]

    for r in rows:
        data.append([
            r.fecha_hora.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_hora else "",
            r.nombreZona or "",
            r.codigo_trabajador or "",
            r.cedula or "",
            f"{r.nombre} {r.apellido}",
            r.correo or "",
            r.telefono or "",
            "CUMPLE" if r.cumple_epp else "NO CUMPLE",
            r.detalle_fallo or "",
            r.observaciones or "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    story.append(table)
    doc.build(story)

    return buffer.getvalue()

# -----------------------------
# 3.4) EXCEL: Asistencia por zona y rango (INSPECTOR)
# -----------------------------
def generar_excel_asistencia(
    db: Session,
    id_inspector: int,
    id_zona: int,
    fecha_desde: str,
    fecha_hasta: str,
) -> bytes:
    d = _parse_date(fecha_desde)
    h = _end_of_day(_parse_date(fecha_hasta))

    rows = (
        db.query(
            RegistroAsistencia.fecha_hora,
            Zona.nombreZona,
            Trabajador.codigo_trabajador,
            Persona.cedula,
            Persona.nombre,
            Persona.apellido,
            RegistroAsistencia.cumple_epp,
            EvidenciaFallo.detalle_fallo,
        )
        .join(Zona, Zona.id_Zona == RegistroAsistencia.id_zona)
        .join(Trabajador, Trabajador.id_trabajador == RegistroAsistencia.id_trabajador)
        .join(Persona, Persona.id_persona == Trabajador.id_persona_trabajador)
        .outerjoin(EvidenciaFallo, EvidenciaFallo.id_registro == RegistroAsistencia.id_registro)
        .filter(
    RegistroAsistencia.id_inspector == id_inspector,
    RegistroAsistencia.id_zona == id_zona,
    RegistroAsistencia.fecha_hora >= d,
    RegistroAsistencia.fecha_hora <= h,
)

        .order_by(RegistroAsistencia.fecha_hora.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    headers = [
        "Fecha/Hora", "Zona", "Código",
        "Cédula", "Trabajador",
        "Estado EPP", "Detalle"
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r.fecha_hora.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_hora else "",
            r.nombreZona or "",
            r.codigo_trabajador or "",
            r.cedula or "",
            f"{r.nombre} {r.apellido}",
            "CUMPLE" if r.cumple_epp else "NO CUMPLE",
            r.detalle_fallo or "",
        ])

    ws.freeze_panes = "A2"

    for col in range(1, len(headers) + 1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

import json

def registrar_reporte(
    db: Session,
    tipo_reporte: str,
    formato: str,
    filtros: dict | None,
    generado_por: str,
    id_empresa: int,
    id_inspector: int | None = None,
) -> Reporte:
    """
    Registra un reporte en la base de datos.
    """
    try:
        # ✅ IMPORTANTE: Convertir diccionario a JSON string
        filtros_json = json.dumps(filtros, ensure_ascii=False) if filtros else None
        
        reporte = Reporte(
            tipo_reporte=tipo_reporte,
            formato=formato,
            filtros=filtros_json,  # ✅ JSON string, NO diccionario
            generado_por=generado_por,
            id_empresa=id_empresa,
            id_inspector=id_inspector,
            borrado=True,
        )
        
        db.add(reporte)
        db.commit()
        db.refresh(reporte)
        
        return reporte
        
    except Exception as e:
        db.rollback()
        print(f"Error al registrar reporte: {str(e)}")
        return None


def obtener_empresa_por_inspector(
    db: Session, 
    id_inspector: int
) -> Dict[str, Any]:
    """
    Obtiene empresa por ID de inspector.
    
    Flujo: Inspector → RegistroSupervisorInspector → Supervisor → Empresa
    
    Args:
        db: Sesión de base de datos
        id_inspector: ID del inspector
        
    Returns:
        Dict con id_Empresa y nombreEmpresa
        
    Raises:
        HTTPException 404: Si inspector no existe o no tiene supervisor/empresa
    """
    # Query con joins explícitos
    resultado = db.query(
        Empresa.id_Empresa,
        Empresa.nombreEmpresa
    ).join(
        Supervisor, 
        Supervisor.id_empresa_supervisor == Empresa.id_Empresa
    ).join(
        RegistroSupervisorInspector, 
        RegistroSupervisorInspector.id_supervisor_registro == Supervisor.id_supervisor
    ).join(
        Inspector, 
        Inspector.id_inspector == RegistroSupervisorInspector.id_inspector_registro
    ).filter(
        Inspector.id_inspector == id_inspector,
        Inspector.borrado == True,
        Supervisor.borrado == True,
        Empresa.borrado == True,
        RegistroSupervisorInspector.borrado == True
    ).first()
    
    if not resultado:
        raise HTTPException(
            status_code=404,
            detail="Inspector no encontrado o sin supervisor/empresa asignado"
        )
    
    return {
        "id_Empresa": resultado[0],
        "nombreEmpresa": resultado[1]
    }


def obtener_zonas_por_inspector(
    db: Session, 
    id_inspector: int
) -> List[dict]:
    """
    Obtiene las zonas asignadas a un inspector.
    
    Flujo: Inspector → InspectorZona → Zona
    
    Args:
        db: Sesión de base de datos
        id_inspector: ID del inspector
        
    Returns:
        Lista de zonas con id y nombre
    """
    # Query optimizada con join directo
    zonas = db.query(
        Zona.id_Zona,
        Zona.nombreZona
    ).join(
        InspectorZona,
        InspectorZona.id_zona_inspectorzona == Zona.id_Zona
    ).filter(
        InspectorZona.id_inspector_inspectorzona == id_inspector,
        InspectorZona.borrado == True,
        Zona.borrado == True
    ).all()
    
    if not zonas:
        return []
    
    return [
        {
            "id": z[0],
            "nombre": z[1]
        }
        for z in zonas
    ]
